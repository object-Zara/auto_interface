# -*- coding:utf-8 -*-
# @datetime:2019/4/1 18:30
# @author:123
# @email:1111@sina.com
# @File:do_excel.py

from openpyxl import load_workbook
from common import get_path


class Case:
    def __init__(self):
        self.case_id = None
        self.case_excute = None
        self.title = None
        self.url = None
        self.data = None
        self.method = None
        self.expect = None
        self.actual = None
        self.result = None


class DoExcelHttp:
    def __init__(self, file_name):
        self.file_name = file_name
        self.wb = load_workbook(file_name)

    def get_init_data(self, sheet_name):
        """读取默认的手机号"""
        sheet = self.wb[sheet_name]
        mobile = sheet.cell(1, 2).value
        self.wb.close()
        return mobile

    def read_data(self, sheet_name):
        sheet = self.wb[sheet_name]
        list_data = []
        for row in range(2, sheet.max_row+1):
            case = Case()
            case.case_id = sheet.cell(row, 1).value
            case.case_excute = sheet.cell(row, 2).value
            case.title = sheet.cell(row, 3).value
            case.url = sheet.cell(row, 4).value
            case.data = sheet.cell(row, 5).value
            case.method = sheet.cell(row, 6).value
            case.expect = sheet.cell(row, 7).value
            case.sql = sheet.cell(row, 10).value
            list_data.append(case)
        self.wb.close()
        return list_data

    def write_data(self, sheet_name, row, column, value):
        sheet = self.wb[sheet_name]
        sheet.cell(row, column, value)
        self.wb.save(self.file_name)
        self.wb.close()


if __name__ == '__main__':
    doexcelhttp = DoExcelHttp(get_path.GetPath().get_test_case_data_path())
    cases = doexcelhttp.read_data('register')
    for case_item in cases:
        print(case_item.data)
    # print(type(case[1].data))

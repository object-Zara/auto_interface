# -*- coding:utf-8 -*-
# @datetime:2019/4/1 19:37
# @author:Xiaoyuan
# @email:Object_ycm@sina.com
# @File:send_email.py

import smtplib
from email.utils import formataddr
from smtplib import SMTP_SSL
from email.mime.text import MIMEText
from email.header import Header
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from openpyxl import load_workbook
from common import get_path


class SendEMail:
    def __init__(self):
        self.wb = load_workbook(get_path.GetPath().get_email_info_path())
        self.sheet_login = self.wb['sender']
        # 邮箱和服务商配置
        self.sender_email = self.sheet_login.cell(2, 2).value
        self.pwd = self.sheet_login.cell(2, 3).value
        self.host_server = self.sheet_login.cell(2, 1).value
        # self.receiver_email = receiver_email
        self.msg = MIMEMultipart()
        self.wb.close()

    @staticmethod
    def receiver_email():
        wb = load_workbook(get_path.GetPath().get_email_info_path())
        sheet_receive = wb['receiver']
        receive_data = []
        for row in range(2, sheet_receive.max_row+1):
            if sheet_receive.cell(row, 1).value:
                receive_data.append(sheet_receive.cell(row, 1).value)
            else:
                continue
        wb.close()
        return receive_data

    def add_content(self, attach_type, attachment_path, attachment_name):
        """添加附件到MIMEMultipart"""
        with open(attachment_path, 'rb') as f:
            mime = MIMEApplication(f.read(), attach_type)
            mime.add_header('Content-Disposition', 'attachment', filename=attachment_name)
            self.msg.attach(mime)

    def send_email(self, from_nickname: str, subject, text_part):
        try:
            """添加正文到MIMEMultipart"""
            self.msg.attach(MIMEText(text_part, 'plain', 'utf-8'))
            for receiver in self.receiver_email():
                # 邮件主体设置
                self.msg['From'] = formataddr([from_nickname, self.sender_email])
                self.msg['To'] = receiver
                self.msg['Subject'] = Header(subject, 'utf-8')
                # msg['Subject'] = Header(subject, 'utf-8').encode()

                with SMTP_SSL(self.host_server, 465) as server:
                    # server.set_debuglevel(1)  # 打印出和SMTP服务器交互的所有信息
                    server.login(self.sender_email, self.pwd)
                    server.sendmail(self.sender_email, receiver, self.msg.as_string())
                    print('邮件发送成功')
        except smtplib.SMTPException:
            print("Error: 无法发送邮件")


if __name__ == '__main__':
    smail = SendEMail()
    smail.add_content('html', r'../test_result/html_report/2019-04-01测试报告.html', '测试报告.html')
    # smail.add_content('image', 'png', 'conf/表情.png', '表情.png')
    smail.send_email('彩', '测试报告', '接口测试报告')


